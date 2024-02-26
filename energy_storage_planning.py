import os
import time
from copy import copy
import pandas as pd
from math import isclose
import pyomo.opt as po
import pyomo.environ as pe
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from network_data import NetworkData
from planning_parameters import PlanningParameters
from energy_storage_parameters import EnergyStorageParameters
from helper_functions import *


# ======================================================================================================================
#   Class ENERGY STORAGE PLANNING
# ======================================================================================================================
class EnergyStoragePlanning:

    def __init__(self, data_dir, filename):
        self.name = filename.replace('.json', '')
        self.data_dir = data_dir
        self.filename = filename
        self.market_data_file = str()
        self.results_dir = os.path.join(data_dir, 'Results')
        self.plots_dir = os.path.join(data_dir, 'Results', 'Plots')
        self.diagrams_dir = os.path.join(data_dir, 'Diagrams')
        self.params_file = str()
        self.ess_params_file = str()
        self.ess_investment_costs_file = str()
        self.years = dict()
        self.days = dict()
        self.num_instants = int()
        self.discount_factor = float()
        self.candidate_nodes = list()
        self.cost_energy_p = dict()
        self.prob_market_scenarios = list()
        self.network = NetworkData()
        self.params = PlanningParameters()
        self.ess_params = EnergyStorageParameters()
        self.investment_costs = dict()

    def run_planning_problem(self):
        _run_planning_problem(self)

    def build_master_problem_model(self):
        return _build_master_problem_model(self)

    def build_subproblem_model(self, ess_params=dict()):
        return self.network.build_model(ess_params=ess_params)

    def update_admm_consensus_variables(self, master_problem_model, subproblem_models, consensus_vars, dual_vars, consensus_vars_prev_iter):
        return _update_admm_consensus_variables(self, master_problem_model, subproblem_models, consensus_vars, dual_vars, consensus_vars_prev_iter)

    def compute_primal_value(self, tso_model, dso_models):
        return _compute_primal_value(self, tso_model, dso_models)

    def compute_master_problem_primal_value(self, master_problem_model):
        return _compute_master_problem_primal_value(self, master_problem_model)

    def compute_subproblems_primal_value(self, subproblem_models):
        return _compute_subproblems_primal_value(self, subproblem_models)

    def get_candidate_solution(self, master_problem_model):
        years = [year for year in self.years]
        candidate_solution = {'investment': {}, 'total_capacity': {}}
        for y in range(len(self.years)):
            year = years[y]
            candidate_solution['investment'][year] = dict()
            candidate_solution['total_capacity'][year] = dict()
            for e in range(len(self.candidate_nodes)):
                node_id = self.candidate_nodes[e]
                candidate_solution['investment'][year][node_id] = dict()
                candidate_solution['investment'][year][node_id]['s'] = abs(pe.value(master_problem_model.es_s_invesment[e, y]))
                candidate_solution['investment'][year][node_id]['e'] = abs(pe.value(master_problem_model.es_e_invesment[e, y]))
                candidate_solution['total_capacity'][year][node_id] = dict()
                candidate_solution['total_capacity'][year][node_id]['s'] = abs(pe.value(master_problem_model.es_s_rated[e, y]))
                candidate_solution['total_capacity'][year][node_id]['e'] = abs(pe.value(master_problem_model.es_e_rated[e, y]))
        return candidate_solution

    def optimize(self, model, from_warm_start=False):
        return _optimize(model, self.ess_params.solver_params, from_warm_start=from_warm_start)

    def run_operational_planning(self, candidate_solution=dict()):
        print('[INFO] Running OPERATIONAL PLANNING...')
        models = self.network.build_model()
        if not candidate_solution:
            candidate_solution = self.get_initial_candidate_solution()
        self.update_model_with_candidate_solution(models, candidate_solution['total_capacity'])
        results = self.network.run_operational_planning_problem(models)
        self.write_operational_planning_results_to_excel(models, results)
        return results, models

    def get_initial_candidate_solution(self):
        return _get_initial_candidate_solution(self)

    def update_model_with_candidate_solution(self, models, candidate_solution):
        for year in self.years:
            for day in self.days:
                s_base = self.network.network[year][day].baseMVA
                for e in range(len(self.candidate_nodes)):
                    node_id = self.candidate_nodes[e]
                    models[year][day].es_planning_s_rated_fixed[e].fix(candidate_solution[year][node_id]['s'] / s_base)
                    models[year][day].es_planning_e_rated_fixed[e].fix(candidate_solution[year][node_id]['e'] / s_base)

    def read_planning_problem(self):
        _read_planning_problem(self)

    def read_market_data_from_file(self):
        _read_market_data_from_file(self)

    def read_planning_parameters_from_file(self):
        print(f'[INFO] Reading PLANNING PARAMETERS from file {self.params_file} ...')
        filename = os.path.join(self.data_dir, self.params_file)
        self.params.read_parameters_from_file(filename)

    def read_energy_storage_parameters_from_file(self):
        print(f'[INFO] Reading ENERGY STORAGE PARAMETERS from file {self.ess_params_file} ...')
        filename = os.path.join(self.data_dir, self.ess_params_file)
        self.ess_params.read_parameters_from_file(filename)

    def read_energy_storage_investment_costs(self):
        filename = os.path.join(self.data_dir, self.ess_investment_costs_file)
        self.investment_costs = _get_investment_costs_from_excel_file(filename, 'Investment Cost', len(self.years))

    def write_planning_results_to_excel(self, optimization_models, results, bound_evolution):
        filename = os.path.join(self.results_dir, self.name + '_planning_results.xlsx')
        processed_results = self.network.process_results(optimization_models, results)
        _write_planning_results_to_excel(self, processed_results['results'], bound_evolution, filename=filename)

    def write_operational_planning_results_to_excel(self, optimization_models, results):
        filename = os.path.join(self.results_dir, self.name + '_operational_planning_results.xlsx')
        processed_results = self.network.process_results(optimization_models, results)
        _write_operational_planning_results_to_excel(self, processed_results['results'], filename=filename)


# ======================================================================================================================
#  OPTIMIZATION  functions
# ======================================================================================================================
def _run_planning_problem(planning_problem):

    years = [year for year in planning_problem.years]
    days = [day for day in planning_problem.days]
    results = dict()

    # ------------------------------------------------------------------------------------------------------------------
    # 0. Initialization
    print('[INFO]\t - Initializing...')
    start = time.time()
    primal_evolution = list()

    # ------------------------------------------------------------------------------------------------------------------
    # 0. Initialization
    # - ADMM variables
    consensus_vars, dual_vars, consensus_vars_prev_iter = create_admm_variables(planning_problem)

    # - Create subproblem models, warm start, update to ADMM
    subproblem_models = planning_problem.build_subproblem_model(ess_params=planning_problem.ess_params)
    planning_problem.network.run_operational_planning_problem(subproblem_models)    # Warm start
    for e in range(len(planning_problem.candidate_nodes)):
        node_id = planning_problem.candidate_nodes[e]
        for year in years:
            s_rated_year = 0.00
            e_rated_year = 0.00
            for day in days:
                num_days = planning_problem.days[day]
                s_base = planning_problem.network.network[year][day].baseMVA
                s_rated_year += pe.value(subproblem_models[year][day].es_planning_s_rated[e]) * s_base * (num_days / 365.00)
                e_rated_year += pe.value(subproblem_models[year][day].es_planning_e_rated[e]) * s_base * (num_days / 365.00)
            consensus_vars['subproblem'][node_id][year]['s'] = s_rated_year
            consensus_vars['subproblem'][node_id][year]['e'] = e_rated_year
    update_subproblems_to_admm(planning_problem, subproblem_models)

    # - Create master problem models, warm start, update to ADMM
    master_problem_model = planning_problem.build_master_problem_model()
    planning_problem.optimize(master_problem_model)
    for e in range(len(planning_problem.candidate_nodes)):
        node_id = planning_problem.candidate_nodes[e]
        for y in range(len(years)):
            year = years[y]
            s_rated_year = pe.value(master_problem_model.es_s_rated[e, y])
            e_rated_year = pe.value(master_problem_model.es_e_rated[e, y])
            consensus_vars['master_problem'][node_id][year]['s'] = s_rated_year
            consensus_vars['master_problem'][node_id][year]['e'] = e_rated_year
    update_master_problem_to_admm(planning_problem, master_problem_model)

    # ------------------------------------------------------------------------------------------------------------------
    # ADMM -- Main cycle
    # ------------------------------------------------------------------------------------------------------------------
    convergence, num_iter = False, 1
    for iter in range(planning_problem.params.num_max_iters):

        print(f'[INFO]\t - ADMM. Iter {num_iter}...')

        iter_start = time.time()

        results['master_problem'] = update_master_problem_and_solve(planning_problem, master_problem_model, consensus_vars['subproblem'], dual_vars['subproblem'])

        # - Update ADMM consensus variables
        planning_problem.update_admm_consensus_variables(master_problem_model, subproblem_models, consensus_vars, dual_vars, consensus_vars_prev_iter)

        # - Update primal evolution
        primal_evolution.append(planning_problem.compute_primal_value(master_problem_model, subproblem_models))

        # - Stopping criteria evaluation
        if iter > 1:
            convergence = planning_problem.check_admm_convergence(consensus_vars, consensus_vars_prev_iter)
            if convergence:
                break

        # - Solve subproblems
        results['subproblems'] = update_subproblems_and_solve(planning_problem, subproblem_models, consensus_vars['master_problem'], dual_vars['master_problem'])

        # - Update ADMM CONSENSUS variables
        planning_problem.update_admm_consensus_variables(master_problem_model, subproblem_models, consensus_vars, dual_vars, consensus_vars_prev_iter)

        # - Update primal evolution
        primal_evolution.append(planning_problem.compute_primal_value(master_problem_model, subproblem_models))

        # 3.3 STOPPING CRITERIA evaluation
        convergence = planning_problem.check_admm_convergence(consensus_vars, consensus_vars_prev_iter)
        if convergence:
            break

        iter_end = time.time()
        print('[INFO] \t - Iter {}: {:.2f} s'.format(num_iter, iter_end - iter_start))
        num_iter += 1

    if not convergence:
        print(f'[WARNING] ADMM did NOT converge in {planning_problem.params.num_max_iters} iterations!')
    else:
        print(f'[INFO] \t - ADMM converged in {iter + 1} iterations.')

    end = time.time()
    total_execution_time = end - start
    print('[INFO] \t - Total execution time: {:.2f}s.'.format(total_execution_time))


def _update_admm_consensus_variables(planning_problem, master_problem_model, subproblem_models, consensus_vars, dual_vars, consensus_vars_prev_iter):
    _update_previous_consensus_variables(planning_problem, consensus_vars, consensus_vars_prev_iter)
    _update_consensus_variables(planning_problem, master_problem_model, subproblem_models, consensus_vars, dual_vars)


def _update_previous_consensus_variables(planning_problem, consensus_vars, consensus_vars_prev_iter):
    for node_id in planning_problem.candidate_nodes:
        for year in planning_problem.years:
            consensus_vars_prev_iter['master_problem'][node_id][year]['s'] = copy(consensus_vars['master_problem'][node_id][year]['s'])
            consensus_vars_prev_iter['master_problem'][node_id][year]['e'] = copy(consensus_vars['master_problem'][node_id][year]['e'])
            consensus_vars_prev_iter['subproblem'][node_id][year]['s'] = copy(consensus_vars['subproblem'][node_id][year]['s'])
            consensus_vars_prev_iter['subproblem'][node_id][year]['e'] = copy(consensus_vars['subproblem'][node_id][year]['e'])


def _update_consensus_variables(planning_problem, master_problem_model, subproblem_models, consensus_vars, dual_vars):

    years = [year for year in planning_problem.years]

    # - Master problem
    for e in master_problem_model.energy_storages:
        node_id = planning_problem.candidate_nodes[e]
        for y in master_problem_model.years:
            year = years[y]
            consensus_vars['master_problem'][node_id][year]['s'] = pe.value(master_problem_model.es_s_rated[e, y])
            consensus_vars['master_problem'][node_id][year]['e'] = pe.value(master_problem_model.es_e_rated[e, y])

    # - Subproblems
    for year in planning_problem.years:
        for e in range(len(planning_problem.candidate_nodes)):
            node_id = planning_problem.candidate_nodes[e]
            consensus_vars['subproblem'][node_id][year]['s'] = 0.00
            consensus_vars['subproblem'][node_id][year]['e'] = 0.00
            for day in planning_problem.days:
                subproblem_model = subproblem_models[year][day]
                s_base = planning_problem.network.network[year][day].baseMVA
                num_days = planning_problem.days[day]
                consensus_vars['subproblem'][node_id][year]['s'] += pe.value(subproblem_model.es_planning_s_rated[e]) * s_base * (num_days / 365.00)
                consensus_vars['subproblem'][node_id][year]['e'] += pe.value(subproblem_model.es_planning_e_rated[e]) * s_base * (num_days / 365.00)

    # Update Lambdas
    for node_id in planning_problem.candidate_nodes:
        for year in planning_problem.years:

            error_s_rated = consensus_vars['master_problem'][node_id][year]['s'] - consensus_vars['subproblem'][node_id][year]['s']
            error_e_rated = consensus_vars['master_problem'][node_id][year]['e'] - consensus_vars['subproblem'][node_id][year]['e']

            dual_vars['master_problem'][node_id][year]['s'] += planning_problem.params.rho_s * (error_s_rated)
            dual_vars['master_problem'][node_id][year]['e'] += planning_problem.params.rho_e * (error_e_rated)
            dual_vars['master_problem'][node_id][year]['s'] += planning_problem.params.rho_s * (-error_s_rated)
            dual_vars['master_problem'][node_id][year]['e'] += planning_problem.params.rho_e * (-error_e_rated)


def update_subproblems_to_admm(planning_problem, subproblem_models):

    e_max = planning_problem.ess_params.max_capacity
    s_max = e_max * planning_problem.ess_params.max_se_factor

    for year in planning_problem.years:
        for day in planning_problem.days:

            subproblem_model = subproblem_models[year][day]
            init_of_value = pe.value(subproblem_model.objective)

            # Add ADMM variables
            subproblem_model.rho_s = pe.Var(domain=pe.NonNegativeReals)
            subproblem_model.rho_s.fix(planning_problem.params.rho_s)
            subproblem_model.rho_e = pe.Var(domain=pe.NonNegativeReals)
            subproblem_model.rho_e.fix(planning_problem.params.rho_s)

            subproblem_model.es_s_rated_req = pe.Var(subproblem_model.energy_storages_planning, domain=pe.NonNegativeReals)
            subproblem_model.es_e_rated_req = pe.Var(subproblem_model.energy_storages_planning, domain=pe.NonNegativeReals)
            subproblem_model.dual_es_s_rated = pe.Var(subproblem_model.energy_storages_planning, domain=pe.Reals)
            subproblem_model.dual_es_e_rated = pe.Var(subproblem_model.energy_storages_planning, domain=pe.Reals)

            # Objective function - augmented Lagrangian
            obj = subproblem_model.objective.expr / max(abs(init_of_value), 1.00)

            # Augmented Lagrangian -- Srated and Erated (residual balancing)
            for e in subproblem_model.energy_storages_planning:
                constraint_s_req = (subproblem_model.es_planning_s_rated[e] - subproblem_model.es_s_rated_req[e]) / abs(s_max)
                constraint_e_req = (subproblem_model.es_planning_e_rated[e] - subproblem_model.es_e_rated_req[e]) / abs(e_max)
                obj += (subproblem_model.dual_es_s_rated[e]) * (constraint_s_req)
                obj += (subproblem_model.dual_es_e_rated[e]) * (constraint_e_req)
                obj += (subproblem_model.rho_s / 2) * (constraint_s_req) ** 2
                obj += (subproblem_model.rho_e / 2) * (constraint_e_req) ** 2

            subproblem_model.objective.expr = obj


def update_subproblems_and_solve(planning_problem, subproblem_models, ess_req, dual_ess):

    print('[INFO] \t\t - Updating subproblems...')

    for year in planning_problem.years:
        for day in planning_problem.days:
            subproblem_model = subproblem_models[year][day]
            s_base = planning_problem.network.network[year][day].baseMVA
            for e in subproblem_model.energy_storages_planning:
                node_id = planning_problem.candidate_nodes[e]
                subproblem_model.dual_es_s_rated[e].fix(dual_ess[node_id][year]['s'] / s_base)
                subproblem_model.dual_es_e_rated[e].fix(dual_ess[node_id][year]['e'] / s_base)
                subproblem_model.es_s_rated_req[e].fix(ess_req[node_id][year]['s'] / s_base)
                subproblem_model.es_e_rated_req[e].fix(ess_req[node_id][year]['e'] / s_base)

    # Solve!
    res = planning_problem.network.run_operational_planning_problem(subproblem_models)
    return res



def update_master_problem_to_admm(planning_problem, master_problem_model):

    init_of_value = planning_problem.ess_params.budget
    e_max = planning_problem.ess_params.max_capacity
    s_max = e_max * planning_problem.ess_params.max_se_factor

    # Add ADMM variables
    master_problem_model.rho_s = pe.Var(domain=pe.NonNegativeReals)
    master_problem_model.rho_s.fix(planning_problem.params.rho_s)
    master_problem_model.rho_e = pe.Var(domain=pe.NonNegativeReals)
    master_problem_model.rho_e.fix(planning_problem.params.rho_s)

    master_problem_model.es_s_rated_req = pe.Var(master_problem_model.energy_storages, master_problem_model.years, domain=pe.NonNegativeReals)
    master_problem_model.es_e_rated_req = pe.Var(master_problem_model.energy_storages, master_problem_model.years, domain=pe.NonNegativeReals)
    master_problem_model.dual_es_s_rated = pe.Var(master_problem_model.energy_storages, master_problem_model.years, domain=pe.Reals)
    master_problem_model.dual_es_e_rated = pe.Var(master_problem_model.energy_storages, master_problem_model.years, domain=pe.Reals)

    # Objective function - augmented Lagrangian
    obj = master_problem_model.objective.expr / max(abs(init_of_value), 1.00)

    # Augmented Lagrangian -- Srated and Erated (residual balancing)
    for e in master_problem_model.energy_storages:
        for y in master_problem_model.years:
            constraint_s_req = (master_problem_model.es_s_rated[e, y] - master_problem_model.es_s_rated_req[e, y]) / abs(s_max)
            constraint_e_req = (master_problem_model.es_e_rated[e, y] - master_problem_model.es_e_rated_req[e, y]) / abs(e_max)
            obj += (master_problem_model.dual_es_s_rated[e, y]) * (constraint_s_req)
            obj += (master_problem_model.dual_es_e_rated[e, y]) * (constraint_e_req)
            obj += (master_problem_model.rho_s / 2) * (constraint_s_req) ** 2
            obj += (master_problem_model.rho_e / 2) * (constraint_e_req) ** 2

    master_problem_model.objective.expr = obj


def update_master_problem_and_solve(planning_problem, master_problem_model, ess_req, dual_ess):

    print('[INFO] \t\t - Updating master problem...')
    years = [year for year in planning_problem.years]

    for e in master_problem_model.energy_storages:
        node_id = planning_problem.candidate_nodes[e]
        for y in master_problem_model.years:
            year = years[y]
            master_problem_model.dual_es_s_rated[e, y].fix(dual_ess[node_id][year]['s'])
            master_problem_model.dual_es_e_rated[e, y].fix(dual_ess[node_id][year]['e'])
            master_problem_model.es_s_rated_req[e, y].fix(ess_req[node_id][year]['s'])
            master_problem_model.es_e_rated_req[e, y].fix(ess_req[node_id][year]['e'])

    # Solve!
    res = planning_problem.optimize(master_problem_model)
    return res


def _build_master_problem_model(planning_problem):

    years = [year for year in planning_problem.years]
    ess_params = planning_problem.ess_params

    model = pe.ConcreteModel()
    model.name = "ESS Optimization -- Benders' Master Problem"

    # ------------------------------------------------------------------------------------------------------------------
    # Sets
    model.years = range(len(planning_problem.years))
    model.energy_storages = range(len(planning_problem.candidate_nodes))

    # ------------------------------------------------------------------------------------------------------------------
    # Decision variables
    model.es_s_invesment = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.0)   # Investment in power capacity in year y
    model.es_e_invesment = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.0)   # Investment in energy capacity in year y
    model.es_s_rated = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.0)       # Total rated power capacity (considering calendar life)
    model.es_e_rated = pe.Var(model.energy_storages, model.years, domain=pe.NonNegativeReals, initialize=0.0)       # Total rated energy capacity (considering calendar life, not considering degradation)
    model.alpha = pe.Var(domain=pe.Reals)                                                           # alpha (associated with cuts) will try to rebuild y in the original problem
    model.alpha.setlb(-1e12)

    # ------------------------------------------------------------------------------------------------------------------
    # Constraints
    # - Yearly Power and Energy ratings as a function of yearly investments
    model.rated_s_capacity = pe.ConstraintList()
    model.rated_e_capacity = pe.ConstraintList()
    for e in model.energy_storages:
        total_s_capacity_per_year = [0.0 for _ in model.years]
        total_e_capacity_per_year = [0.0 for _ in model.years]
        for y in model.years:
            year = years[y]
            num_years = planning_problem.years[year]
            tcal_norm = round(ess_params.t_cal / num_years)
            max_tcal_norm = min(y + tcal_norm, len(years))
            for x in range(y, max_tcal_norm):
                total_s_capacity_per_year[x] += model.es_s_invesment[e, y]
                total_e_capacity_per_year[x] += model.es_e_invesment[e, y]
        for y in model.years:
            model.rated_s_capacity.add(model.es_s_rated[e, y] == total_s_capacity_per_year[y])
            model.rated_e_capacity.add(model.es_e_rated[e, y] == total_e_capacity_per_year[y])

    # - Maximum Energy Capacity (related to space constraints)
    model.energy_storage_maximum_capacity = pe.ConstraintList()
    for e in model.energy_storages:
        for y in model.years:
            model.energy_storage_maximum_capacity.add(model.es_e_rated[e, y] <= ess_params.max_capacity)

    # - S/E factor
    model.energy_storage_power_to_energy_factor = pe.ConstraintList()
    for e in model.energy_storages:
        for y in model.years:
            model.energy_storage_power_to_energy_factor.add(model.es_s_rated[e, y] >= model.es_e_rated[e, y] * ess_params.min_se_factor)
            model.energy_storage_power_to_energy_factor.add(model.es_s_rated[e, y] <= model.es_e_rated[e, y] * ess_params.max_se_factor)

    # - Maximum Investment Cost
    investment_cost_total = 0.0
    model.energy_storage_investment = pe.ConstraintList()
    for y in model.years:
        year = years[y]
        c_inv_s = planning_problem.investment_costs['power_capacity'][year]
        c_inv_e = planning_problem.investment_costs['energy_capacity'][year]
        annualization = 1 / ((1 + planning_problem.discount_factor) ** (int(year) - int(years[0])))
        for e in model.energy_storages:
            investment_cost_total += annualization * model.es_s_invesment[e, y] * c_inv_s
            investment_cost_total += annualization * model.es_e_invesment[e, y] * c_inv_e
    model.energy_storage_investment.add(investment_cost_total <= ess_params.budget)

    # Benders' cuts
    model.benders_cuts = pe.ConstraintList()

    # Objective function
    investment_cost = 0.0
    for e in model.energy_storages:
        for y in model.years:
            year = years[y]
            c_inv_s = planning_problem.investment_costs['power_capacity'][year]
            c_inv_e = planning_problem.investment_costs['energy_capacity'][year]
            annualization = 1 / ((1 + planning_problem.discount_factor) ** (int(year) - int(years[0])))

            # Investment Cost
            investment_cost += annualization * model.es_s_invesment[e, y] * c_inv_s
            investment_cost += annualization * model.es_e_invesment[e, y] * c_inv_e

    obj = investment_cost + model.alpha
    model.objective = pe.Objective(sense=pe.minimize, expr=obj)

    # Model suffixes (used for warm start)
    model.ipopt_zL_out = pe.Suffix(direction=pe.Suffix.IMPORT)  # Ipopt bound multipliers (obtained from solution)
    model.ipopt_zU_out = pe.Suffix(direction=pe.Suffix.IMPORT)
    model.ipopt_zL_in = pe.Suffix(direction=pe.Suffix.EXPORT)  # Ipopt bound multipliers (sent to solver)
    model.ipopt_zU_in = pe.Suffix(direction=pe.Suffix.EXPORT)
    model.dual = pe.Suffix(direction=pe.Suffix.IMPORT_EXPORT)  # Obtain dual solutions from previous solve and send to warm start

    return model


def _compute_primal_value(planning_problem, master_problem_model, subproblem_models):
    primal_value = 0.00
    primal_value += planning_problem.compute_master_problem_primal_value(master_problem_model)
    primal_value += planning_problem.compute_subproblems_primal_value(subproblem_models)
    return primal_value


def _compute_master_problem_primal_value(planning_problem, master_problem_model):
    years = [year for year in planning_problem.years]
    primal_value = 0.00
    for e in master_problem_model.energy_storages:
        for y in master_problem_model.years:
            year = years[y]
            c_inv_s = planning_problem.investment_costs['power_capacity'][year]
            c_inv_e = planning_problem.investment_costs['energy_capacity'][year]
            annualization = 1 / ((1 + planning_problem.discount_factor) ** (int(year) - int(years[0])))
            primal_value += annualization * pe.value(master_problem_model.es_s_invesment[e, y]) * c_inv_s
            primal_value += annualization * pe.value(master_problem_model.es_e_invesment[e, y]) * c_inv_e
    return primal_value


def _compute_subproblems_primal_value(planning_problem, subproblem_models):
    primal_value = planning_problem.network.get_objective_function_value(subproblem_models)
    return primal_value


def _optimize(model, params, from_warm_start=False):

    solver = po.SolverFactory(params.solver, executable=params.solver_path, tee=params.verbose)

    if from_warm_start:
        model.ipopt_zL_in.update(model.ipopt_zL_out)
        model.ipopt_zU_in.update(model.ipopt_zU_out)
        solver.options['warm_start_init_point'] = 'yes'
        solver.options['warm_start_bound_push'] = 1e-9
        solver.options['warm_start_mult_bound_push'] = 1e-9
        solver.options['mu_init'] = 1e-9

    if params.verbose:
        solver.options['print_level'] = 6
        solver.options['output_file'] = 'optim_log.txt'

    if params.solver == 'ipopt':
        solver.options['tol'] = params.solver_tol
        solver.options['acceptable_tol'] = params.solver_tol * 1e3
        solver.options['acceptable_iter'] = 5
        solver.options['nlp_scaling_method'] = 'none'
        solver.options['max_iter'] = 10000
        solver.options['linear_solver'] = params.linear_solver

    result = solver.solve(model, tee=params.verbose)
    '''
    if not result.solver.status == po.SolverStatus.ok:
        import logging
        from pyomo.util.infeasible import log_infeasible_constraints
        filename = os.path.join(os.getcwd(), 'master_problem.log')
        print(log_infeasible_constraints(model, log_expression=True, log_variables=True))
        logging.basicConfig(filename=filename, encoding='utf-8', level=logging.INFO)
    '''

    return result


def create_admm_variables(planning_problem):

    consensus_variables = {
        'master_problem': dict(),
        'subproblem': dict()
    }

    dual_variables = {
        'master_problem': dict(),
        'subproblem': dict()
    }

    consensus_variables_prev_iter = {
        'master_problem': dict(),
        'subproblem': dict()
    }

    for node_id in planning_problem.candidate_nodes:

        consensus_variables['master_problem'][node_id] = dict()
        consensus_variables['subproblem'][node_id] = dict()
        dual_variables['master_problem'][node_id] = dict()
        dual_variables['subproblem'][node_id] = dict()
        consensus_variables_prev_iter['master_problem'][node_id] = dict()
        consensus_variables_prev_iter['subproblem'][node_id] = dict()

        for year in planning_problem.years:
            consensus_variables['master_problem'][node_id][year] = {'s': 0.00, 'e': 0.00}
            consensus_variables['subproblem'][node_id][year] = {'s': 0.00, 'e': 0.00}
            dual_variables['master_problem'][node_id][year] = {'s': 0.00, 'e': 0.00}
            dual_variables['subproblem'][node_id][year] = {'s': 0.00, 'e': 0.00}
            consensus_variables_prev_iter['master_problem'][node_id][year] = {'s': 0.00, 'e': 0.00}
            consensus_variables_prev_iter['subproblem'][node_id][year] = {'s': 0.00, 'e': 0.00}

    return consensus_variables, dual_variables, consensus_variables_prev_iter


# ======================================================================================================================
#  PLANNING PROBLEM read functions
# ======================================================================================================================
def _read_planning_problem(planning_problem):

    # Create results folder
    if not os.path.exists(planning_problem.results_dir):
        os.makedirs(planning_problem.results_dir)

    # Read specification file
    filename = os.path.join(planning_problem.data_dir, planning_problem.filename)
    planning_data = convert_json_to_dict(read_json_file(filename))

    # General Parameters
    for year in planning_data['Years']:
        planning_problem.years[int(year)] = planning_data['Years'][year]
    planning_problem.days = planning_data['Days']
    planning_problem.num_instants = planning_data['NumInstants']
    planning_problem.candidate_nodes = [int(candidate_node) for candidate_node in planning_data['CandidateNodes']]

    # Market Data
    planning_problem.discount_factor = planning_data['DiscountFactor']
    planning_problem.market_data_file = planning_data['MarketData']
    planning_problem.read_market_data_from_file()

    # Network Data
    planning_problem.network = NetworkData()
    planning_problem.network.name = planning_data['Network']['name']
    planning_problem.network.data_dir = planning_problem.data_dir
    planning_problem.network.results_dir = planning_problem.results_dir
    planning_problem.network.plots_dir = planning_problem.plots_dir
    planning_problem.network.diagrams_dir = planning_problem.diagrams_dir
    planning_problem.network.years = planning_problem.years
    planning_problem.network.days = planning_problem.days
    planning_problem.network.num_instants = planning_problem.num_instants
    planning_problem.network.discount_factor = planning_problem.discount_factor
    planning_problem.network.candidate_nodes = planning_problem.candidate_nodes
    planning_problem.network.prob_market_scenarios = planning_problem.prob_market_scenarios
    planning_problem.network.cost_energy_p = planning_problem.cost_energy_p
    planning_problem.network.params_file = planning_data['Network']['params_file']
    planning_problem.network.read_network_parameters()
    if planning_problem.network.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
        planning_problem.network.prob_market_scenarios = [1.00]
    planning_problem.network.read_network_data()

    # Planning Parameters
    planning_problem.params_file = planning_data['PlanningParameters']['params_file']
    planning_problem.read_planning_parameters_from_file()

    # Energy Storage Parameters
    planning_problem.ess_params_file = planning_data['EnergyStorageParameters']['params_file']
    planning_problem.read_energy_storage_parameters_from_file()

    # Investment Costs
    planning_problem.ess_investment_costs_file = planning_data['EnergyStorageInvestmentCosts']['costs_file']
    planning_problem.read_energy_storage_investment_costs()


def _read_market_data_from_file(planning_problem):
    try:
        for year in planning_problem.years:
            filename = os.path.join(planning_problem.data_dir, 'Market Data', f'{planning_problem.market_data_file}_{year}.xlsx')
            num_scenarios, prob_scenarios = _get_market_scenarios_info_from_excel_file(filename, 'Scenarios')
            planning_problem.prob_market_scenarios = prob_scenarios
            planning_problem.cost_energy_p[year] = dict()
            for day in planning_problem.days:
                planning_problem.cost_energy_p[year][day] = _get_market_costs_from_excel_file(filename, f'Cp, {day}', num_scenarios)
    except:
        print(f'[ERROR] Reading market data from file(s). Exiting...')
        exit(ERROR_SPECIFICATION_FILE)


def _get_market_scenarios_info_from_excel_file(filename, sheet_name):

    num_scenarios = 0
    prob_scenarios = list()

    try:
        df = pd.read_excel(filename, sheet_name=sheet_name, header=None)
        if is_int(df.iloc[0, 1]):
            num_scenarios = int(df.iloc[0, 1])
        for i in range(num_scenarios):
            if is_number(df.iloc[0, i + 2]):
                prob_scenarios.append(float(df.iloc[0, i + 2]))
    except:
        print('[ERROR] Workbook {}. Sheet {} does not exist.'.format(filename, sheet_name))
        exit(1)

    if num_scenarios != len(prob_scenarios):
        print('[WARNING] EnergyStorage file. Number of scenarios different from the probability vector!')

    if round(sum(prob_scenarios), 2) != 1.00:
        print('[ERROR] Probability of scenarios does not add up to 100%. Check file {}. Exiting.'.format(filename))
        exit(ERROR_MARKET_DATA_FILE)

    return num_scenarios, prob_scenarios


def _get_market_costs_from_excel_file(filename, sheet_name, num_scenarios):
    data = pd.read_excel(filename, sheet_name=sheet_name)
    _, num_cols = data.shape
    cost_values = dict()
    scn_idx = 0
    for i in range(num_scenarios):
        cost_values_scenario = list()
        for j in range(num_cols - 1):
            cost_values_scenario.append(float(data.iloc[i, j + 1]))
        cost_values[scn_idx] = cost_values_scenario
        scn_idx = scn_idx + 1
    return cost_values


def _get_investment_costs_from_excel_file(filename, sheet_name, num_years):

    try:

        df = pd.read_excel(filename, sheet_name=sheet_name, header=None)
        data = {
            'power_capacity': dict(),
            'energy_capacity': dict()
        }

        for i in range(num_years):

            year = int(df.iloc[0, i + 1])

            if is_number(df.iloc[1, i + 1]):
                data['power_capacity'][year] = float(df.iloc[1, i + 1])

            if is_number(df.iloc[2, i + 1]):
                data['energy_capacity'][year] = float(df.iloc[2, i + 1])

        return data

    except:
        print('[ERROR] Workbook {}. Sheet {} does not exist.'.format(filename, sheet_name))
        exit(ERROR_MARKET_DATA_FILE)


# ======================================================================================================================
#  RESULTS - write functions
# ======================================================================================================================s
def _write_planning_results_to_excel(planning_problem, results, bound_evolution, filename='operation_planning_results'):

    wb = Workbook()

    _write_operational_planning_main_info_to_excel(planning_problem, wb, results)
    _write_ess_planning_specifications(planning_problem, wb, results)
    _write_bound_evolution_to_excel(wb, bound_evolution)

    # Energy Storages (Planning) results
    _write_energy_storages_planning_results_to_excel(planning_problem, wb, results)

    _write_network_voltage_results_to_excel(planning_problem, wb, results)
    _write_network_consumption_results_to_excel(planning_problem, wb, results)
    _write_network_generation_results_to_excel(planning_problem, wb, results)
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'losses')
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'ratio')
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'current_perc')
    _write_network_branch_power_flow_results_to_excel(planning_problem, wb, results)
    _write_network_energy_storages_results_to_excel(planning_problem, wb, results)

    # Save results
    try:
        wb.save(filename)
    except:
        from datetime import datetime
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = f"{filename.replace('xlsx', '')}_{current_time}.xlsx"
        print(f"[WARNING] Results saved to file {backup_filename}.xlsx")
        wb.save(backup_filename)


def _write_operational_planning_results_to_excel(planning_problem, results, filename='operation_planning_results'):

    wb = Workbook()

    _write_operational_planning_main_info_to_excel(planning_problem, wb, results)
    _write_ess_planning_specifications(planning_problem, wb, results)

    # Energy Storages (Planning) results
    _write_energy_storages_planning_results_to_excel(planning_problem, wb, results)

    _write_network_voltage_results_to_excel(planning_problem, wb, results)
    _write_network_consumption_results_to_excel(planning_problem, wb, results)
    _write_network_generation_results_to_excel(planning_problem, wb, results)
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'losses')
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'ratio')
    _write_network_branch_results_to_excel(planning_problem, wb, results, 'current_perc')
    _write_network_branch_power_flow_results_to_excel(planning_problem, wb, results)
    _write_network_energy_storages_results_to_excel(planning_problem, wb, results)

    # Save results
    try:
        wb.save(filename)
    except:
        from datetime import datetime
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d_%H-%M-%S")
        backup_filename = f"{filename.replace('xlsx', '')}_{current_time}.xlsx"
        print(f"[WARNING] Results saved to file {backup_filename}.xlsx")
        wb.save(backup_filename)


def _write_bound_evolution_to_excel(workbook, bound_evolution):

    sheet = workbook.create_sheet('Convergence Characteristic')

    lower_bound = bound_evolution['lower_bound']
    upper_bound = bound_evolution['upper_bound']
    num_lines = max(len(upper_bound), len(lower_bound))

    num_style = '0.00'

    # Write header
    line_idx = 1
    sheet.cell(row=line_idx, column=1).value = 'Iteration'
    sheet.cell(row=line_idx, column=2).value = 'Lower Bound, [NPV Mm.u.]'
    sheet.cell(row=line_idx, column=3).value = 'Upper Bound, [NPV Mm.u.]'

    # Iterations
    line_idx = 2
    for i in range(num_lines):
        sheet.cell(row=line_idx, column=1).value = i
        line_idx += 1

    # Lower bound
    line_idx = 2
    for value in lower_bound:
        sheet.cell(row=line_idx, column=2).value = value / 1e6
        sheet.cell(row=line_idx, column=2).number_format = num_style
        line_idx += 1

    # Upper bound
    line_idx = 2
    for value in upper_bound:
        sheet.cell(row=line_idx, column=3).value = value / 1e6
        sheet.cell(row=line_idx, column=3).number_format = num_style
        line_idx += 1


def _write_operational_planning_main_info_to_excel(planning_problem, workbook, results):

    network = planning_problem.network

    sheet = workbook.worksheets[0]
    sheet.title = 'Main Info'
    decimal_style = '0.00'
    line_idx = 1

    # Write Header
    col_idx = 2
    for year in planning_problem.years:
        for _ in planning_problem.days:
            sheet.cell(row=line_idx, column=col_idx).value = year
            col_idx += 1

    line_idx += 1
    col_idx = 2
    for _ in planning_problem.years:
        for day in planning_problem.days:
            sheet.cell(row=line_idx, column=col_idx).value = day
            col_idx += 1

    # - Objective
    line_idx += 1
    col_idx = 1
    obj_string = 'Objective'
    if network.params.obj_type == OBJ_MIN_COST:
        obj_string += ' (cost), [€]'
    elif network.params.obj_type == OBJ_CONGESTION_MANAGEMENT:
        obj_string += ' (congestion management)'
    sheet.cell(row=line_idx, column=col_idx).value = obj_string
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['obj']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Total Load
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Load, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            load_aux = results[year][day]['total_load']
            if network.params.l_curt:
                load_aux -= results[year][day]['load_curt']
            sheet.cell(row=line_idx, column=col_idx).value = load_aux
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Flexibility used
    if network.params.fl_reg:
        line_idx += 1
        col_idx = 1
        sheet.cell(row=line_idx, column=col_idx).value = 'Flexibility used, [MWh]'
        col_idx += 1
        for year in results:
            for day in results[year]:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['flex_used']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1

    # Total Load curtailed
    if network.params.l_curt:
        line_idx += 1
        col_idx = 1
        sheet.cell(row=line_idx, column=col_idx).value = 'Load curtailed, [MWh]'
        col_idx += 1
        for year in results:
            for day in results[year]:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['load_curt']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1

    # Total Generation
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Generation, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_gen']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Total Conventional Generation
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Conventional Generation, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_conventional_gen']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Total Renewable Generation
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Renewable generation, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['total_renewable_gen']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Renewable Generation Curtailed
    if network.params.rg_curt:
        line_idx += 1
        col_idx = 1
        sheet.cell(row=line_idx, column=col_idx).value = 'Renewable generation curtailed, [MWh]'
        col_idx += 1
        for year in results:
            for day in results[year]:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['gen_curt']
                sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
                col_idx += 1

    # Losses
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Losses, [MWh]'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['losses']
            sheet.cell(row=line_idx, column=col_idx).number_format = decimal_style
            col_idx += 1

    # Number of price (market) scenarios
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Number of market scenarios'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = len(network.network[year][day].prob_market_scenarios)
            col_idx += 1

    # Number of operation (generation and consumption) scenarios
    line_idx += 1
    col_idx = 1
    sheet.cell(row=line_idx, column=col_idx).value = 'Number of operation scenarios'
    col_idx += 1
    for year in results:
        for day in results[year]:
            sheet.cell(row=line_idx, column=col_idx).value = len(network.network[year][day].prob_operation_scenarios)
            col_idx += 1


def _write_ess_planning_specifications(planning_problem, workbook, results):

    sheet = workbook.create_sheet('ESS Planning, Specifications')

    years = [year for year in planning_problem.years]
    days = [day for day in planning_problem.days]

    num_style = '0.00'

    # Write Header
    line_idx = 1
    col_idx = 3
    for year in years:
        for _ in days:
            sheet.cell(row=line_idx, column=col_idx).value = year
            col_idx += 1

    line_idx += 1
    sheet.cell(row=line_idx, column=1).value = 'Node'
    sheet.cell(row=line_idx, column=2).value = 'Quantity'
    col_idx = 3
    for _ in years:
        for day in days:
            sheet.cell(row=line_idx, column=col_idx).value = day
            col_idx += 1

    # Write investment values, power and energy
    for node_id in planning_problem.candidate_nodes:

        # Power capacity
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'S, [MVA]'
        col_idx = 3
        for year in years:
            for day in days:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['scenarios'][0][0]['energy_storages_planning']['s_rated'][node_id]
                sheet.cell(row=line_idx, column=col_idx).number_format = num_style
                col_idx += 1

        # Energy capacity
        line_idx = line_idx + 1
        sheet.cell(row=line_idx, column=1).value = node_id
        sheet.cell(row=line_idx, column=2).value = 'E, [MVAh]'
        col_idx = 3
        for year in years:
            for day in days:
                sheet.cell(row=line_idx, column=col_idx).value = results[year][day]['scenarios'][0][0]['energy_storages_planning']['e_rated'][node_id]
                sheet.cell(row=line_idx, column=col_idx).number_format = num_style
                col_idx += 1


def _write_energy_storages_planning_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Energy Storage (Planning)')

    row_idx = 1
    decimal_style = '0.00'
    percent_style = '0.00%'

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 7).value = p

    for year in results:
        for day in results[year]:

            expected_p = dict()
            expected_q = dict()
            expected_s = dict()
            expected_soc = dict()
            expected_soc_percent = dict()
            for node_id in planning_problem.candidate_nodes:
                expected_p[node_id] = [0.0 for _ in range(planning_problem.num_instants)]
                expected_q[node_id] = [0.0 for _ in range(planning_problem.num_instants)]
                expected_s[node_id] = [0.0 for _ in range(planning_problem.num_instants)]
                expected_soc[node_id] = [0.0 for _ in range(planning_problem.num_instants)]
                expected_soc_percent[node_id] = [0.0 for _ in range(planning_problem.num_instants)]

            for s_m in results[year][day]['scenarios']:

                omega_m = planning_problem.network.network[year][day].prob_market_scenarios[s_m]

                for s_o in results[year][day]['scenarios'][s_m]:

                    omega_s = planning_problem.network.network[year][day].prob_operation_scenarios[s_o]

                    for node_id in planning_problem.candidate_nodes:

                        # Active power
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_p = results[year][day]['scenarios'][s_m][s_o]['energy_storages_planning']['p'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = ess_p
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if ess_p != 'N/A':
                                expected_p[node_id][p] += ess_p * omega_m * omega_s
                            else:
                                expected_p[node_id][p] = ess_p

                        # Reactive power
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_q = results[year][day]['scenarios'][s_m][s_o]['energy_storages_planning']['q'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = ess_q
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if ess_q != 'N/A':
                                expected_q[node_id][p] += ess_q * omega_m * omega_s
                            else:
                                expected_q[node_id][p] = ess_q

                        # Apparent power
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_s = results[year][day]['scenarios'][s_m][s_o]['energy_storages_planning']['s'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = ess_s
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if ess_s != 'N/A':
                                expected_s[node_id][p] += ess_s * omega_m * omega_s
                            else:
                                expected_s[node_id][p] = ess_s

                        # State-of-Charge, [MVAh]
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'SoC, [MVAh]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_soc = results[year][day]['scenarios'][s_m][s_o]['energy_storages_planning']['soc'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = ess_soc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if ess_soc != 'N/A':
                                expected_soc[node_id][p] += ess_soc * omega_m * omega_s
                            else:
                                expected_soc[node_id][p] = ess_soc

                        # State-of-Charge, [%]
                        row_idx = row_idx + 1
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'SoC, [%]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(planning_problem.num_instants):
                            ess_soc_percent = results[year][day]['scenarios'][s_m][s_o]['energy_storages_planning']['soc_percent'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = ess_soc_percent
                            sheet.cell(row=row_idx, column=p + 7).number_format = percent_style
                            if ess_soc_percent != 'N/A':
                                expected_soc_percent[node_id][p] += ess_soc_percent * omega_m * omega_s
                            else:
                                expected_soc_percent[node_id][p] = ess_soc_percent

            for node_id in planning_problem.candidate_nodes:

                # Active Power, [MW]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_p[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style

                # Reactive Power, [MVAr]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_q[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style

                # Apparent Power, [MVA]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_s[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style

                # State-of-Charge, [MVAh]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'SoC, [MVAh]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_soc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style

                # State-of-Charge, [%]
                row_idx = row_idx + 1
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'SoC, [%]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(planning_problem.num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_soc_percent[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = percent_style


def _write_network_voltage_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Voltage')
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    decimal_style = '0.00'

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 7).value = p
    row_idx = row_idx + 1

    # Write results -- TSO
    network = planning_problem.network.network

    for year in results:
        for day in results[year]:

            ref_node_id = network[year][day].get_reference_node_id()
            expected_vmag = dict()
            expected_vang = dict()
            for node in network[year][day].nodes:
                expected_vmag[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_vang[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network[year][day].prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network[year][day].prob_operation_scenarios[s_o]
                    for node_id in results[year][day]['scenarios'][s_m][s_o]['voltage']['vmag']:

                        v_min, v_max = network[year][day].get_node_voltage_limits(node_id)

                        # Voltage magnitude
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Vmag, [p.u.]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network[year][day].num_instants):
                            v_mag = results[year][day]['scenarios'][s_m][s_o]['voltage']['vmag'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = v_mag
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if node_id != ref_node_id and (v_mag > v_max + SMALL_TOLERANCE or v_mag < v_min - SMALL_TOLERANCE):
                                sheet.cell(row=row_idx, column=p + 7).fill = violation_fill
                            expected_vmag[node_id][p] += v_mag * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Voltage angle
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Vang, [º]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network[year][day].num_instants):
                            v_ang = results[year][day]['scenarios'][s_m][s_o]['voltage']['vang'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = v_ang
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            expected_vang[node_id][p] += v_ang * omega_m * omega_s
                        row_idx = row_idx + 1

            for node in network[year][day].nodes:

                node_id = node.bus_i
                v_min, v_max = network[year][day].get_node_voltage_limits(node_id)

                # Expected voltage magnitude
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Vmag, [p.u.]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_vmag[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    if node_id != ref_node_id and (expected_vmag[node_id][p] > v_max + SMALL_TOLERANCE or expected_vmag[node_id][p] < v_min - SMALL_TOLERANCE):
                        sheet.cell(row=row_idx, column=p + 1).fill = violation_fill
                row_idx = row_idx + 1

                # Expected voltage angle
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Vang, [º]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_vang[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1


def _write_network_consumption_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Consumption')
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    decimal_style = '0.00'

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 7).value = p
    row_idx = row_idx + 1

    network = planning_problem.network.network
    params = planning_problem.network.params

    for year in results:
        for day in results[year]:

            expected_pc = dict()
            expected_flex_up = dict()
            expected_flex_down = dict()
            expected_pc_curt = dict()
            expected_pnet = dict()
            expected_qc = dict()
            for node in network[year][day].nodes:
                expected_pc[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_flex_up[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_flex_down[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_pc_curt[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_pnet[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_qc[node.bus_i] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network[year][day].prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network[year][day].prob_operation_scenarios[s_o]
                    for node_id in results[year][day]['scenarios'][s_m][s_o]['consumption']['pc']:

                        # - Active Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Pc, [MW]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network[year][day].num_instants):
                            pc = results[year][day]['scenarios'][s_m][s_o]['consumption']['pc'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = pc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            expected_pc[node_id][p] += pc * omega_m * omega_s
                        row_idx = row_idx + 1

                        if params.fl_reg:

                            # - Flexibility, up
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Flex Up, [MW]'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network[year][day].num_instants):
                                flex = results[year][day]['scenarios'][s_m][s_o]['consumption']['p_up'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = flex
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                expected_flex_up[node_id][p] += flex * omega_m * omega_s
                            row_idx = row_idx + 1

                            # - Flexibility, down
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Flex Down, [MW]'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network[year][day].num_instants):
                                flex = results[year][day]['scenarios'][s_m][s_o]['consumption']['p_down'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = flex
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                expected_flex_down[node_id][p] += flex * omega_m * omega_s
                            row_idx = row_idx + 1

                        if params.l_curt:

                            # - Active power curtailment
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Pc_curt, [MW]'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network[year][day].num_instants):
                                pc_curt = results[year][day]['scenarios'][s_m][s_o]['consumption']['pc_curt'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = pc_curt
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                if pc_curt >= SMALL_TOLERANCE:
                                    sheet.cell(row=row_idx, column=p + 7).fill = violation_fill
                                expected_pc_curt[node_id][p] += pc_curt * omega_m * omega_s
                            row_idx = row_idx + 1

                        if params.fl_reg or params.l_curt:

                            # - Active power net consumption
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = int(year)
                            sheet.cell(row=row_idx, column=3).value = day
                            sheet.cell(row=row_idx, column=4).value = 'Pc_net, [MW]'
                            sheet.cell(row=row_idx, column=5).value = s_m
                            sheet.cell(row=row_idx, column=6).value = s_o
                            for p in range(network[year][day].num_instants):
                                p_net = results[year][day]['scenarios'][s_m][s_o]['consumption']['pc_net'][node_id][p]
                                sheet.cell(row=row_idx, column=p + 7).value = p_net
                                sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                                expected_pnet[node_id][p] += p_net * omega_m * omega_s
                            row_idx = row_idx + 1

                        # - Reactive power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Qc, [MVAr]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network[year][day].num_instants):
                            qc = results[year][day]['scenarios'][s_m][s_o]['consumption']['qc'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = qc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            expected_qc[node_id][p] += qc * omega_m * omega_s
                        row_idx = row_idx + 1

            for node in network[year][day].nodes:

                node_id = node.bus_i

                # - Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Pc, [MW]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_pc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                if params.fl_reg:

                    # - Flexibility, up
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Flex Up, [MW]'
                    sheet.cell(row=row_idx, column=5).value = 'Expected'
                    sheet.cell(row=row_idx, column=6).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 7).value = expected_flex_up[node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1

                    # - Flexibility, down
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Flex Down, [MW]'
                    sheet.cell(row=row_idx, column=5).value = 'Expected'
                    sheet.cell(row=row_idx, column=6).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 7).value = expected_flex_down[node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1

                if params.l_curt:

                    # - Load curtailment (active power)
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Pc_curt, [MW]'
                    sheet.cell(row=row_idx, column=5).value = 'Expected'
                    sheet.cell(row=row_idx, column=6).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 7).value = expected_pc_curt[node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                        if expected_pc_curt[node_id][p] >= SMALL_TOLERANCE:
                            sheet.cell(row=row_idx, column=p + 7).fill = violation_fill
                    row_idx = row_idx + 1

                if params.fl_reg or params.l_curt:

                    # - Active power net consumption
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = int(year)
                    sheet.cell(row=row_idx, column=3).value = day
                    sheet.cell(row=row_idx, column=4).value = 'Pc_net, [MW]'
                    sheet.cell(row=row_idx, column=5).value = 'Expected'
                    sheet.cell(row=row_idx, column=6).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 7).value = expected_pnet[node_id][p]
                        sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                    row_idx = row_idx + 1

                # - Reactive power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Qc, [MVAr]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_qc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1


def _write_network_generation_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Generation')
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    decimal_style = '0.00'

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Generator ID'
    sheet.cell(row=row_idx, column=3).value = 'Type'
    sheet.cell(row=row_idx, column=4).value = 'Year'
    sheet.cell(row=row_idx, column=5).value = 'Day'
    sheet.cell(row=row_idx, column=6).value = 'Quantity'
    sheet.cell(row=row_idx, column=7).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=8).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 9).value = p
    row_idx = row_idx + 1

    network = planning_problem.network.network
    params = planning_problem.network.params

    for year in results:
        for day in results[year]:

            expected_pg = dict()
            expected_pg_curt = dict()
            expected_pg_net = dict()
            expected_qg = dict()
            for generator in network[year][day].generators:
                expected_pg[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_pg_curt[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_pg_net[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_qg[generator.gen_id] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network[year][day].prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network[year][day].prob_operation_scenarios[s_o]
                    for g in results[year][day]['scenarios'][s_m][s_o]['generation']['pg']:

                        node_id = network[year][day].generators[g].bus
                        gen_id = network[year][day].generators[g].gen_id
                        gen_type = network[year][day].get_gen_type(gen_id)

                        # Active Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = gen_id
                        sheet.cell(row=row_idx, column=3).value = gen_type
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'Pg, [MW]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network[year][day].num_instants):
                            pg = results[year][day]['scenarios'][s_m][s_o]['generation']['pg'][g][p]
                            sheet.cell(row=row_idx, column=p + 9).value = pg
                            sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            expected_pg[gen_id][p] += pg * omega_m * omega_s
                        row_idx = row_idx + 1

                        if params.rg_curt:

                            # Active Power curtailment
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = gen_id
                            sheet.cell(row=row_idx, column=3).value = gen_type
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Pg_curt, [MW]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network[year][day].num_instants):
                                pg_curt = results[year][day]['scenarios'][s_m][s_o]['generation']['pg_curt'][g][p]
                                sheet.cell(row=row_idx, column=p + 9).value = pg_curt
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                if pg_curt > SMALL_TOLERANCE:
                                    sheet.cell(row=row_idx, column=p + 9).fill = violation_fill
                                expected_pg_curt[gen_id][p] += pg_curt * omega_m * omega_s
                            row_idx = row_idx + 1

                            # Active Power net
                            sheet.cell(row=row_idx, column=1).value = node_id
                            sheet.cell(row=row_idx, column=2).value = gen_id
                            sheet.cell(row=row_idx, column=3).value = gen_type
                            sheet.cell(row=row_idx, column=4).value = int(year)
                            sheet.cell(row=row_idx, column=5).value = day
                            sheet.cell(row=row_idx, column=6).value = 'Pg_net, [MW]'
                            sheet.cell(row=row_idx, column=7).value = s_m
                            sheet.cell(row=row_idx, column=8).value = s_o
                            for p in range(network[year][day].num_instants):
                                pg_net = results[year][day]['scenarios'][s_m][s_o]['generation']['pg_net'][g][p]
                                sheet.cell(row=row_idx, column=p + 9).value = pg_net
                                sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                                expected_pg_net[gen_id][p] += pg_net * omega_m * omega_s
                            row_idx = row_idx + 1

                        # Reactive Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = gen_id
                        sheet.cell(row=row_idx, column=3).value = gen_type
                        sheet.cell(row=row_idx, column=4).value = int(year)
                        sheet.cell(row=row_idx, column=5).value = day
                        sheet.cell(row=row_idx, column=6).value = 'Qg, [MVAr]'
                        sheet.cell(row=row_idx, column=7).value = s_m
                        sheet.cell(row=row_idx, column=8).value = s_o
                        for p in range(network[year][day].num_instants):
                            qg = results[year][day]['scenarios'][s_m][s_o]['generation']['qg'][g][p]
                            sheet.cell(row=row_idx, column=p + 9).value = qg
                            sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                            expected_qg[gen_id][p] += qg * omega_m * omega_s
                        row_idx = row_idx + 1

            for generator in network[year][day].generators:

                node_id = generator.bus
                gen_id = generator.gen_id
                gen_type = network[year][day].get_gen_type(gen_id)

                # Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = gen_id
                sheet.cell(row=row_idx, column=3).value = gen_type
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Pg, [MW]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_pg[gen_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1

                if params.rg_curt:

                    # Active Power curtailment
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = gen_id
                    sheet.cell(row=row_idx, column=3).value = gen_type
                    sheet.cell(row=row_idx, column=4).value = int(year)
                    sheet.cell(row=row_idx, column=5).value = day
                    sheet.cell(row=row_idx, column=6).value = 'Pg_curt, [MW]'
                    sheet.cell(row=row_idx, column=7).value = 'Expected'
                    sheet.cell(row=row_idx, column=8).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 9).value = expected_pg_curt[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                        if expected_pg_curt[gen_id][p] > SMALL_TOLERANCE:
                            sheet.cell(row=row_idx, column=p + 9).fill = violation_fill
                    row_idx = row_idx + 1

                    # Active Power net
                    sheet.cell(row=row_idx, column=1).value = node_id
                    sheet.cell(row=row_idx, column=2).value = gen_id
                    sheet.cell(row=row_idx, column=3).value = gen_type
                    sheet.cell(row=row_idx, column=4).value = int(year)
                    sheet.cell(row=row_idx, column=5).value = day
                    sheet.cell(row=row_idx, column=6).value = 'Pg_net, [MW]'
                    sheet.cell(row=row_idx, column=7).value = 'Expected'
                    sheet.cell(row=row_idx, column=8).value = '-'
                    for p in range(network[year][day].num_instants):
                        sheet.cell(row=row_idx, column=p + 9).value = expected_pg_net[gen_id][p]
                        sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                    row_idx = row_idx + 1

                # Reactive Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = gen_id
                sheet.cell(row=row_idx, column=3).value = gen_type
                sheet.cell(row=row_idx, column=4).value = int(year)
                sheet.cell(row=row_idx, column=5).value = day
                sheet.cell(row=row_idx, column=6).value = 'Qg, [MVAr]'
                sheet.cell(row=row_idx, column=7).value = 'Expected'
                sheet.cell(row=row_idx, column=8).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 9).value = expected_qg[gen_id][p]
                    sheet.cell(row=row_idx, column=p + 9).number_format = decimal_style
                row_idx = row_idx + 1


def _write_network_branch_results_to_excel(planning_problem, workbook, results, result_type):

    sheet_name = str()
    if result_type == 'losses':
        sheet_name = 'Branch Losses'
    elif result_type == 'ratio':
        sheet_name = 'Transformer Ratio'
    elif result_type == 'current_perc':
        sheet_name = 'Branch Loading'
    sheet = workbook.create_sheet(sheet_name)
    violation_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    decimal_style = '0.00'
    perc_style = '0.00%'

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'From Node ID'
    sheet.cell(row=row_idx, column=2).value = 'To Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Year'
    sheet.cell(row=row_idx, column=4).value = 'Day'
    sheet.cell(row=row_idx, column=5).value = 'Quantity'
    sheet.cell(row=row_idx, column=6).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=7).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 8).value = p
    row_idx = row_idx + 1

    network = planning_problem.network.network

    aux_string = str()
    if result_type == 'losses':
        aux_string = 'P, [MW]'
    elif result_type == 'ratio':
        aux_string = 'Ratio'
    elif result_type == 'current_perc':
        aux_string = 'I, [%]'

    for year in results:
        for day in results[year]:

            expected_values = dict()
            for k in range(len(network[year][day].branches)):
                expected_values[k] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network[year][day].prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network[year][day].prob_operation_scenarios[s_o]
                    for k in results[year][day]['scenarios'][s_m][s_o]['branches'][result_type]:
                        branch = network[year][day].branches[k]
                        if not(result_type == 'ratio' and not branch.is_transformer):

                            sheet.cell(row=row_idx, column=1).value = branch.fbus
                            sheet.cell(row=row_idx, column=2).value = branch.tbus
                            sheet.cell(row=row_idx, column=3).value = int(year)
                            sheet.cell(row=row_idx, column=4).value = day
                            sheet.cell(row=row_idx, column=5).value = aux_string
                            sheet.cell(row=row_idx, column=6).value = s_m
                            sheet.cell(row=row_idx, column=7).value = s_o
                            for p in range(network[year][day].num_instants):
                                value = results[year][day]['scenarios'][s_m][s_o]['branches'][result_type][k][p]
                                if result_type == 'current_perc':
                                    sheet.cell(row=row_idx, column=p + 8).value = value
                                    sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                                    if value > 1.0 + SMALL_TOLERANCE:
                                        sheet.cell(row=row_idx, column=p + 8).fill = violation_fill
                                else:
                                    sheet.cell(row=row_idx, column=p + 8).value = value
                                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                                expected_values[k][p] += value * omega_m * omega_s
                            row_idx = row_idx + 1

            for k in range(len(network[year][day].branches)):
                branch = network[year][day].branches[k]
                if not (result_type == 'ratio' and not branch.is_transformer):

                    sheet.cell(row=row_idx, column=1).value = branch.fbus
                    sheet.cell(row=row_idx, column=2).value = branch.tbus
                    sheet.cell(row=row_idx, column=3).value = int(year)
                    sheet.cell(row=row_idx, column=4).value = day
                    sheet.cell(row=row_idx, column=5).value = aux_string
                    sheet.cell(row=row_idx, column=6).value = 'Expected'
                    sheet.cell(row=row_idx, column=7).value = '-'
                    for p in range(network[year][day].num_instants):
                        if result_type == 'current_perc':
                            sheet.cell(row=row_idx, column=p + 8).value = expected_values[k][p]
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                            if expected_values[k][p] > 1.0:
                                sheet.cell(row=row_idx, column=p + 8).fill = violation_fill
                        else:
                            sheet.cell(row=row_idx, column=p + 8).value = expected_values[k][p]
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                    row_idx = row_idx + 1


def _write_network_branch_power_flow_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Power Flows')
    decimal_style = '0.00'
    perc_style = '0.00%'

    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'From Node ID'
    sheet.cell(row=row_idx, column=2).value = 'To Node ID'
    sheet.cell(row=row_idx, column=3).value = 'Year'
    sheet.cell(row=row_idx, column=4).value = 'Day'
    sheet.cell(row=row_idx, column=5).value = 'Quantity'
    sheet.cell(row=row_idx, column=6).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=7).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 8).value = p
    row_idx = row_idx + 1

    network = planning_problem.network.network

    for year in results:
        for day in results[year]:

            expected_values = {'pij': {}, 'pji': {}, 'qij': {}, 'qji': {}, 'sij': {}, 'sji': {}}
            for k in range(len(network[year][day].branches)):
                expected_values['pij'][k] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['pji'][k] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['qij'][k] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['qji'][k] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['sij'][k] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_values['sji'][k] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network[year][day].prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network[year][day].prob_operation_scenarios[s_o]
                    for k in range(len(network[year][day].branches)):

                        branch = network[year][day].branches[k]
                        rating = branch.rate
                        if rating == 0.0:
                            rating = BRANCH_UNKNOWN_RATING

                        # Pij, [MW]
                        sheet.cell(row=row_idx, column=1).value = branch.fbus
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pij'][k][p]
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_values['pij'][k][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Pij, [%]
                        sheet.cell(row=row_idx, column=1).value = branch.fbus
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pij'][k][p] / rating)
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                        row_idx = row_idx + 1

                        # Pji, [MW]
                        sheet.cell(row=row_idx, column=1).value = branch.tbus
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pji'][k][p]
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_values['pji'][k][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Pji, [%]
                        sheet.cell(row=row_idx, column=1).value = branch.tbus
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'P, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['pji'][k][p] / rating)
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                        row_idx = row_idx + 1

                        # Qij, [MVAr]
                        sheet.cell(row=row_idx, column=1).value = branch.fbus
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qij'][k][p]
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_values['qij'][k][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Qij, [%]
                        sheet.cell(row=row_idx, column=1).value = branch.fbus
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qij'][k][p] / rating)
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                        row_idx = row_idx + 1

                        # Qji, [MW]
                        sheet.cell(row=row_idx, column=1).value = branch.tbus
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qji'][k][p]
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_values['qji'][k][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Qji, [%]
                        sheet.cell(row=row_idx, column=1).value = branch.tbus
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'Q, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['qji'][k][p] / rating)
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                        row_idx = row_idx + 1

                        # Sij, [MVA]
                        sheet.cell(row=row_idx, column=1).value = branch.fbus
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sij'][k][p]
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_values['sij'][k][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Sij, [%]
                        sheet.cell(row=row_idx, column=1).value = branch.fbus
                        sheet.cell(row=row_idx, column=2).value = branch.tbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'S, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sij'][k][p] / rating)
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                        row_idx = row_idx + 1

                        # Sji, [MW]
                        sheet.cell(row=row_idx, column=1).value = branch.tbus
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sji'][k][p]
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_values['sji'][k][p] += value * omega_m * omega_s
                        row_idx = row_idx + 1

                        # Sji, [%]
                        sheet.cell(row=row_idx, column=1).value = branch.tbus
                        sheet.cell(row=row_idx, column=2).value = branch.fbus
                        sheet.cell(row=row_idx, column=3).value = int(year)
                        sheet.cell(row=row_idx, column=4).value = day
                        sheet.cell(row=row_idx, column=5).value = 'S, [%]'
                        sheet.cell(row=row_idx, column=6).value = s_m
                        sheet.cell(row=row_idx, column=7).value = s_o
                        for p in range(network[year][day].num_instants):
                            value = abs(results[year][day]['scenarios'][s_m][s_o]['branches']['power_flow']['sji'][k][p] / rating)
                            sheet.cell(row=row_idx, column=p + 8).value = value
                            sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                        row_idx = row_idx + 1

            for k in range(len(network[year][day].branches)):

                branch = network[year][day].branches[k]
                rating = branch.rate
                if rating == 0.0:
                    rating = BRANCH_UNKNOWN_RATING

                # Pij, [MW]
                sheet.cell(row=row_idx, column=1).value = branch.fbus
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_values['pij'][k][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # Pij, [%]
                sheet.cell(row=row_idx, column=1).value = branch.fbus
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = abs(expected_values['pij'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                row_idx = row_idx + 1

                # Pji, [MW]
                sheet.cell(row=row_idx, column=1).value = branch.tbus
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_values['pji'][k][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # Pji, [%]
                sheet.cell(row=row_idx, column=1).value = branch.tbus
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'P, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = abs(expected_values['pji'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                row_idx = row_idx + 1

                # Qij, [MVAr]
                sheet.cell(row=row_idx, column=1).value = branch.fbus
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_values['qij'][k][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # Qij, [%]
                sheet.cell(row=row_idx, column=1).value = branch.fbus
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = abs(expected_values['qij'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                row_idx = row_idx + 1

                # Qji, [MVAr]
                sheet.cell(row=row_idx, column=1).value = branch.tbus
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_values['qji'][k][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # Qji, [%]
                sheet.cell(row=row_idx, column=1).value = branch.tbus
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'Q, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = abs(expected_values['qji'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # Sij, [MVA]
                sheet.cell(row=row_idx, column=1).value = branch.fbus
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_values['sij'][k][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # Sij, [%]
                sheet.cell(row=row_idx, column=1).value = branch.fbus
                sheet.cell(row=row_idx, column=2).value = branch.tbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'S, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = abs(expected_values['sij'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                row_idx = row_idx + 1

                # Sji, [MVA]
                sheet.cell(row=row_idx, column=1).value = branch.tbus
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = expected_values['sji'][k][p]
                    sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                row_idx = row_idx + 1

                # Sji, [%]
                sheet.cell(row=row_idx, column=1).value = branch.tbus
                sheet.cell(row=row_idx, column=2).value = branch.fbus
                sheet.cell(row=row_idx, column=3).value = int(year)
                sheet.cell(row=row_idx, column=4).value = day
                sheet.cell(row=row_idx, column=5).value = 'S, [%]'
                sheet.cell(row=row_idx, column=6).value = 'Expected'
                sheet.cell(row=row_idx, column=7).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 8).value = abs(expected_values['sji'][k][p]) / rating
                    sheet.cell(row=row_idx, column=p + 8).number_format = perc_style
                row_idx = row_idx + 1


def _write_network_energy_storages_results_to_excel(planning_problem, workbook, results):

    sheet = workbook.create_sheet('Energy Storage')
    decimal_style = '0.00'
    percent_style = '0.00%'
    row_idx = 1

    # Write Header
    sheet.cell(row=row_idx, column=1).value = 'Network Node ID'
    sheet.cell(row=row_idx, column=2).value = 'Year'
    sheet.cell(row=row_idx, column=3).value = 'Day'
    sheet.cell(row=row_idx, column=4).value = 'Quantity'
    sheet.cell(row=row_idx, column=5).value = 'Market Scenario'
    sheet.cell(row=row_idx, column=6).value = 'Operation Scenario'
    for p in range(planning_problem.num_instants):
        sheet.cell(row=row_idx, column=p + 7).value = p
    row_idx = row_idx + 1

    network = planning_problem.network.network

    for year in results:
        for day in results[year]:

            expected_p = dict()
            expected_q = dict()
            expected_s = dict()
            expected_soc = dict()
            expected_soc_percent = dict()
            for energy_storage in network[year][day].energy_storages:
                expected_p[energy_storage.bus] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_q[energy_storage.bus] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_s[energy_storage.bus] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_soc[energy_storage.bus] = [0.0 for _ in range(network[year][day].num_instants)]
                expected_soc_percent[energy_storage.bus] = [0.0 for _ in range(network[year][day].num_instants)]

            for s_m in results[year][day]['scenarios']:
                omega_m = network[year][day].prob_market_scenarios[s_m]
                for s_o in results[year][day]['scenarios'][s_m]:
                    omega_s = network[year][day].prob_operation_scenarios[s_o]
                    for node_id in results[year][day]['scenarios'][s_m][s_o]['energy_storages']['p']:

                        # - Active Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network[year][day].num_instants):
                            ess_p = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['p'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = ess_p
                            sheet.cell(row=row_idx, column=p + 8).number_format = decimal_style
                            expected_p[node_id][p] += ess_p * omega_m * omega_s
                        row_idx = row_idx + 1

                        # - Reactive Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'Q, [MVAr]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network[year][day].num_instants):
                            ess_q = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['q'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = ess_q
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            expected_q[node_id][p] += ess_q * omega_m * omega_s
                        row_idx = row_idx + 1

                        # - Apparent Power
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'S, [MVA]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network[year][day].num_instants):
                            ess_s = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['s'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = ess_s
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            expected_s[node_id][p] += ess_s * omega_m * omega_s
                        row_idx = row_idx + 1

                        # State-of-Charge, [MWh]
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'SoC, [MWh]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network[year][day].num_instants):
                            ess_soc = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['soc'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = ess_soc
                            sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                            if ess_soc != 'N/A':
                                expected_soc[node_id][p] += ess_soc * omega_m * omega_s
                            else:
                                expected_soc[node_id][p] = ess_soc
                        row_idx = row_idx + 1

                        # State-of-Charge, [%]
                        sheet.cell(row=row_idx, column=1).value = node_id
                        sheet.cell(row=row_idx, column=2).value = int(year)
                        sheet.cell(row=row_idx, column=3).value = day
                        sheet.cell(row=row_idx, column=4).value = 'SoC, [%]'
                        sheet.cell(row=row_idx, column=5).value = s_m
                        sheet.cell(row=row_idx, column=6).value = s_o
                        for p in range(network[year][day].num_instants):
                            ess_soc_percent = results[year][day]['scenarios'][s_m][s_o]['energy_storages']['soc_percent'][node_id][p]
                            sheet.cell(row=row_idx, column=p + 7).value = ess_soc_percent
                            sheet.cell(row=row_idx, column=p + 7).number_format = percent_style
                            if ess_soc_percent != 'N/A':
                                expected_soc_percent[node_id][p] += ess_soc_percent * omega_m * omega_s
                            else:
                                expected_soc_percent[node_id][p] = ess_soc_percent
                        row_idx = row_idx + 1

            for energy_storage in network[year][day].energy_storages:

                node_id = energy_storage.bus

                # - Active Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'P, [MW]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_p[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - Reactive Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'Q, [MVAr]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_q[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # - Apparent Power
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'S, [MVA]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_s[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # State-of-Charge, [MWh]
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'SoC, [MWh]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_soc[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = decimal_style
                row_idx = row_idx + 1

                # State-of-Charge, [%]
                sheet.cell(row=row_idx, column=1).value = node_id
                sheet.cell(row=row_idx, column=2).value = int(year)
                sheet.cell(row=row_idx, column=3).value = day
                sheet.cell(row=row_idx, column=4).value = 'SoC, [%]'
                sheet.cell(row=row_idx, column=5).value = 'Expected'
                sheet.cell(row=row_idx, column=6).value = '-'
                for p in range(network[year][day].num_instants):
                    sheet.cell(row=row_idx, column=p + 7).value = expected_soc_percent[node_id][p]
                    sheet.cell(row=row_idx, column=p + 7).number_format = percent_style
                row_idx = row_idx + 1


# ======================================================================================================================
#   Aux functions
# ======================================================================================================================
def _get_initial_candidate_solution(planning_problem):
    candidate_solution = {'investment': {}, 'total_capacity': {}}
    for year in planning_problem.years:
        candidate_solution['investment'][year] = dict()
        candidate_solution['total_capacity'][year] = dict()
        for node_id in planning_problem.candidate_nodes:
            candidate_solution['investment'][year][node_id] = dict()
            candidate_solution['investment'][year][node_id]['s'] = 0.00
            candidate_solution['investment'][year][node_id]['e'] = 0.00
            candidate_solution['total_capacity'][year][node_id] = dict()
            candidate_solution['total_capacity'][year][node_id]['s'] = 0.00
            candidate_solution['total_capacity'][year][node_id]['e'] = 0.00
    return candidate_solution


def _print_candidate_solution(candidate_solution):

    print('[INFO] Candidate solution:')

    # Header
    print('\t\t{:3}\t{:10}\t'.format('', 'Capacity'), end='')
    for node_id in candidate_solution['total_capacity']:
        for year in candidate_solution['total_capacity'][node_id]:
            print(f'{year}\t\t', end='')
        print()
        break

    # Values
    for node_id in candidate_solution['total_capacity']:
        print('\t\t{:3}\t{:10}\t'.format(node_id, 'S, [MVA]'), end='')
        for year in candidate_solution['total_capacity'][node_id]:
            print("{:.3f}\t".format(candidate_solution['total_capacity'][node_id][year]['s']), end='')
        print()
        print('\t\t{:3}\t{:10}\t'.format(node_id, 'E, [MVAh]'), end='')
        for year in candidate_solution['total_capacity'][node_id]:
            print("{:.3f}\t".format(candidate_solution['total_capacity'][node_id][year]['e']), end='')
        print()
